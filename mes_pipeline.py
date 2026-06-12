# -*- coding: utf-8 -*-
"""매달 실행하는 통합 파이프라인.
사용법: python mes_pipeline.py <패트롤_데이터.xlsx> [출력_접두어]

동작: 입력 데이터로 (1)지식베이스 재학습 → (2)각 기록을 '그 시점 이전 이력' 기준으로 채점
      → (3)원본 컬럼 그대로 + 등급/점수/사유 3컬럼을 붙여 원래 순서로 출력.
출력: <접두어>_MES업로드.csv  (MES 인입용, 원본순서 유지)
      <접두어>_검토용.xlsx     (긴급도 색상, 점수순 정렬)
      knowledge_base.json       (재학습 결과)
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import patrol_lib as P
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
# MES에 내보낼 등급 코드 (필요시 현장 코드체계에 맞게 수정)
GRADE_CODE = {'긴급': 1, '주의': 2, '일반': 3}

def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, '스텐 공정패트롤활동 데이터(260612).xlsx')
    if not os.path.isfile(inp):
        inp = r'C:\Users\Admin\Downloads\스텐 공정패트롤활동 데이터(260612).xlsx'
    pref = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(os.path.basename(inp))[0]

    raw = pd.read_excel(inp)
    raw.columns = [str(c).strip() for c in raw.columns]
    orig_cols = list(raw.columns)
    raw['_row'] = range(len(raw))  # 원본 순서 보존용

    # (1) 재학습
    kb = P.train_kb(raw)
    with open(os.path.join(HERE, 'knowledge_base.json'), 'w', encoding='utf-8') as f:
        json.dump(kb, f, ensure_ascii=False, indent=1)

    # (2) 시점기준 채점
    scored = P.score(raw, kb)

    # (3) 원본순서 복원 + 3컬럼 부착
    scored['등급코드'] = scored['긴급도'].map(GRADE_CODE)
    scored = scored.sort_values('_row')
    out = raw[['_row'] + orig_cols].merge(
        scored[['_row', '긴급도', '등급코드', '우선순위점수', '우선순위사유']], on='_row'
    ).drop(columns='_row')
    out = out.rename(columns={'긴급도': '등급', '우선순위점수': '점수', '우선순위사유': '사유'})

    # MES 업로드용 CSV (원본순서, Excel 한글깨짐 방지 utf-8-sig)
    csv_path = os.path.join(HERE, f'{pref}_MES업로드.csv')
    out.to_csv(csv_path, index=False, encoding='utf-8-sig')

    # 검토용 색상 엑셀 (점수순 정렬)
    xlsx_path = os.path.join(HERE, f'{pref}_검토용.xlsx')
    review = out.sort_values('점수', ascending=False)
    fills = {'긴급': 'F4C0C0', '주의': 'FBE2B7', '일반': 'E5EFD9'}
    fonts = {'긴급': '901010', '주의': '7A4B00', '일반': '3B5A1E'}
    thin = Side(style='thin', color='D0D0D0')
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as w:
        review.to_excel(w, sheet_name='우선순위', index=False)
        ws = w.book['우선순위']
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color='2F5B7C')
            c.alignment = Alignment(horizontal='center', vertical='center')
        gcol = list(out.columns).index('등급')
        for row in ws.iter_rows(min_row=2):
            grade = row[gcol].value
            fill = PatternFill('solid', start_color=fills.get(grade, 'FFFFFF'))
            for c in row:
                c.fill = fill
                c.font = Font(name='Arial', size=10, color=fonts.get(grade, '000000'), bold=(grade == '긴급'))
                c.border = Border(bottom=thin)
        for col in ws.columns:
            L = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(L * 1.6 + 2, 48)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    n = out['등급'].value_counts()
    print(f"재학습 {kb['n_train']}건 ({kb['train_period']})")
    print(f"채점 {len(out)}건 → 긴급 {n.get('긴급',0)} / 주의 {n.get('주의',0)} / 일반 {n.get('일반',0)}")
    print(f"MES 업로드: {os.path.basename(csv_path)}")
    print(f"검토용:     {os.path.basename(xlsx_path)}")

if __name__ == '__main__':
    main()
