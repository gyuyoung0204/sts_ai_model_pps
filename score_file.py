# -*- coding: utf-8 -*-
"""사용법: python score_file.py <새_패트롤_엑셀.xlsx> [출력.xlsx]
지식베이스(knowledge_base.json)를 이용해 새 데이터에 우선순위를 매기고
긴급도 색상이 입혀진 엑셀을 생성한다."""
import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import patrol_lib as P
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, '_샘플입력.xlsx')
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(inp)[0] + '_우선순위.xlsx'
    with open(os.path.join(HERE, 'knowledge_base.json'), encoding='utf-8') as f:
        kb = json.load(f)

    df = pd.read_excel(inp)
    scored = P.score(df, kb)

    cols = ['긴급도', '우선순위점수', '불량처리 일시', '공정', '설비', '표준대분류', '표준세분류',
            '위치유형', '위치번호', '심각도', '관련부품', '우선순위사유', '불량내용', '검사원']
    res = scored[[c for c in cols if c in scored.columns]].copy()

    fills = {'긴급': 'F4C0C0', '주의': 'FBE2B7', '일반': 'E5EFD9'}
    fonts = {'긴급': '901010', '주의': '7A4B00', '일반': '3B5A1E'}
    thin = Side(style='thin', color='D0D0D0')

    with pd.ExcelWriter(out, engine='openpyxl') as w:
        res.to_excel(w, sheet_name='우선순위', index=False)
        ws = w.book['우선순위']
        # 헤더
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color='2F5B7C')
            c.alignment = Alignment(horizontal='center', vertical='center')
        # 행 색상 (긴급도 기준)
        for row in ws.iter_rows(min_row=2):
            grade = row[0].value
            fill = PatternFill('solid', start_color=fills.get(grade, 'FFFFFF'))
            for c in row:
                c.fill = fill
                c.font = Font(name='Arial', size=10, color=fonts.get(grade, '000000'),
                              bold=(grade == '긴급'))
                c.border = Border(bottom=thin)
                c.alignment = Alignment(vertical='center', wrap_text=(c.column_letter in ('L',)))
        # 폭
        widths = {'A':7,'B':10,'C':17,'D':10,'E':17,'F':12,'G':14,'H':8,'I':7,'J':8,'K':14,'L':46,'M':30,'N':8}
        for col, wd in widths.items():
            ws.column_dimensions[col].width = wd
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    n = scored['긴급도'].value_counts()
    print(f"채점 완료: 총 {len(scored)}건 → 긴급 {n.get('긴급',0)} / 주의 {n.get('주의',0)} / 일반 {n.get('일반',0)}")
    print(f"저장: {out}")
    print(f"(학습기준: {kb['train_period']}, {kb['n_train']}건)")

if __name__ == '__main__':
    main()
