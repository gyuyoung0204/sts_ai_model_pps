# -*- coding: utf-8 -*-
"""[2단계] 예측 — 학습된 모델로 '그달 새로 발생한 데이터'에 등급/코드/점수/사유를 매긴다.
사용법: python predict_new.py <그달_새데이터.xlsx> [출력접두어]
* 새 데이터는 학습에 쓰지 않음. knowledge_base.json(이미 학습된 모델)만 불러 예측.
출력: <접두어>_예측결과.csv  (원본순서 유지, MES 인입용)
      <접두어>_예측결과.xlsx (등급 색상, 점수순)
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import patrol_api as API
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))

def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, '_샘플입력_6월.xlsx')
    pref = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(os.path.basename(inp))[0]

    kb = API.load_model()                       # 학습된 모델 로드
    raw = pd.read_excel(inp)
    raw.columns = [str(c).strip() for c in raw.columns]

    records = raw.to_dict('records')            # DB 행처럼 레코드화
    scored = API.score_records(records, kb=kb)  # 모델 호출 → 4필드 예측
    out = pd.DataFrame(scored)

    csv_path = os.path.join(HERE, f'{pref}_예측결과.csv')
    out.to_csv(csv_path, index=False, encoding='utf-8-sig')

    # 검토용 색상 엑셀
    xlsx_path = os.path.join(HERE, f'{pref}_예측결과.xlsx')
    review = out.sort_values(['등급코드', '점수'], ascending=[True, False])
    fills = {'긴급': 'F4C0C0', '주의': 'FBE2B7', '일반': 'E5EFD9'}
    fonts = {'긴급': '901010', '주의': '7A4B00', '일반': '3B5A1E'}
    thin = Side(style='thin', color='D0D0D0')
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as w:
        review.to_excel(w, sheet_name='예측결과', index=False)
        ws = w.book['예측결과']
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color='2F5B7C')
            c.alignment = Alignment(horizontal='center', vertical='center')
        gcol = list(out.columns).index('등급')
        for row in ws.iter_rows(min_row=2):
            grade = row[gcol].value
            fill = PatternFill('solid', start_color=fills.get(grade, 'FFFFFF'))
            for c in row:
                c.fill = fill
                c.font = Font(name='Arial', size=10, color=fonts.get(grade, '000000'), bold=(grade == '긴급'))
                c.border = Border(bottom=thin)
        for col in ws.columns:
            L = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(L * 1.6 + 2, 48)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    n = out['등급'].value_counts()
    print(f"[예측 완료] 학습모델 기준 {kb['train_period']} ({kb['n_train']}건 학습)")
    print(f"  새 데이터 {len(out)}건 → 긴급 {n.get('긴급',0)} / 주의 {n.get('주의',0)} / 일반 {n.get('일반',0)}")
    print(f"  MES 인입용: {os.path.basename(csv_path)}")
    print(f"  검토용:     {os.path.basename(xlsx_path)}")

if __name__ == '__main__':
    main()
